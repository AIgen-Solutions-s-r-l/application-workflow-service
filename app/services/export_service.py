"""
Export service for generating CSV and Excel exports of application data.

Provides:
- CSV export of applications
- Excel export with formatting
- Streaming exports for large datasets
"""

import csv
import io
from collections.abc import AsyncGenerator
from datetime import datetime

from app.core.mongo import failed_applications_collection, success_applications_collection
from app.log.logging import logger


class ExportService:
    """
    Service for exporting application data to various formats.
    """

    # Fields to export (in order)
    EXPORT_FIELDS = [
        "application_id",
        "portal",
        "title",
        "company_name",
        "location",
        "status",
        "created_at",
        "applied_at",
        "error_reason",
    ]

    FIELD_LABELS = {
        "application_id": "Application ID",
        "portal": "Portal",
        "title": "Job Title",
        "company_name": "Company",
        "location": "Location",
        "status": "Status",
        "created_at": "Created At",
        "applied_at": "Applied At",
        "error_reason": "Error Reason",
    }

    async def export_to_csv(
        self,
        user_id: str,
        include_successful: bool = True,
        include_failed: bool = True,
        portal_filter: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> str:
        """
        Export applications to CSV format.

        Args:
            user_id: The user ID to export data for.
            include_successful: Include successful applications.
            include_failed: Include failed applications.
            portal_filter: Optional portal filter.
            date_from: Optional start date filter.
            date_to: Optional end date filter.

        Returns:
            CSV string content.
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        headers = [self.FIELD_LABELS.get(f, f) for f in self.EXPORT_FIELDS]
        writer.writerow(headers)

        # Fetch and write data
        if include_successful:
            async for row in self._fetch_applications(
                user_id=user_id,
                collection=success_applications_collection,
                status="success",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                writer.writerow(row)

        if include_failed:
            async for row in self._fetch_applications(
                user_id=user_id,
                collection=failed_applications_collection,
                status="failed",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                writer.writerow(row)

        return output.getvalue()

    async def export_to_csv_stream(
        self,
        user_id: str,
        include_successful: bool = True,
        include_failed: bool = True,
        portal_filter: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> AsyncGenerator[str, None]:
        """
        Stream applications as CSV for large datasets.

        Args:
            user_id: The user ID to export data for.
            include_successful: Include successful applications.
            include_failed: Include failed applications.
            portal_filter: Optional portal filter.
            date_from: Optional start date filter.
            date_to: Optional end date filter.

        Yields:
            CSV rows as strings.
        """
        # Yield header
        output = io.StringIO()
        writer = csv.writer(output)
        headers = [self.FIELD_LABELS.get(f, f) for f in self.EXPORT_FIELDS]
        writer.writerow(headers)
        yield output.getvalue()

        # Yield data rows
        if include_successful:
            async for row in self._fetch_applications(
                user_id=user_id,
                collection=success_applications_collection,
                status="success",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(row)
                yield output.getvalue()

        if include_failed:
            async for row in self._fetch_applications(
                user_id=user_id,
                collection=failed_applications_collection,
                status="failed",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(row)
                yield output.getvalue()

    async def export_to_excel(
        self,
        user_id: str,
        include_successful: bool = True,
        include_failed: bool = True,
        portal_filter: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> bytes:
        """
        Export applications to Excel format.

        Args:
            user_id: The user ID to export data for.
            include_successful: Include successful applications.
            include_failed: Include failed applications.
            portal_filter: Optional portal filter.
            date_from: Optional start date filter.
            date_to: Optional end date filter.

        Returns:
            Excel file as bytes.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            csv_content = await self.export_to_csv(
                user_id=user_id,
                include_successful=include_successful,
                include_failed=include_failed,
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            )
            return csv_content.encode("utf-8")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Write header
        headers = [self.FIELD_LABELS.get(f, f) for f in self.EXPORT_FIELDS]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        row_num = 2

        if include_successful:
            async for row_data in self._fetch_applications(
                user_id=user_id,
                collection=success_applications_collection,
                status="success",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.fill = success_fill
                row_num += 1

        if include_failed:
            async for row_data in self._fetch_applications(
                user_id=user_id,
                collection=failed_applications_collection,
                status="failed",
                portal_filter=portal_filter,
                date_from=date_from,
                date_to=date_to,
            ):
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.fill = failed_fill
                row_num += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = max(len(str(cell.value or "")) for cell in ws[column_letter])
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    async def _fetch_applications(
        self,
        user_id: str,
        collection,
        status: str,
        portal_filter: str | None = None,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> AsyncGenerator[list, None]:
        """
        Fetch applications from a collection and yield rows.

        Args:
            user_id: The user ID.
            collection: MongoDB collection to query.
            status: Status label for these applications.
            portal_filter: Optional portal filter.
            date_from: Optional start date.
            date_to: Optional end date.

        Yields:
            List of field values for each application.
        """
        doc = await collection.find_one({"user_id": user_id})

        if not doc or "content" not in doc:
            return

        for app_id, job_data in doc.get("content", {}).items():
            # Apply portal filter
            if portal_filter:
                portal = job_data.get("portal", "")
                if portal.lower() != portal_filter.lower():
                    continue

            # Apply date filters
            created_at = job_data.get("created_at") or job_data.get("applied_at")
            if created_at:
                if isinstance(created_at, str):
                    try:
                        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                    except ValueError:
                        created_at = None

                if created_at:
                    if date_from and created_at < date_from:
                        continue
                    if date_to and created_at > date_to:
                        continue

            # Build row
            row = []
            for field in self.EXPORT_FIELDS:
                if field == "application_id":
                    value = app_id
                elif field == "status":
                    value = status
                else:
                    value = job_data.get(field, "")

                # Format datetime fields
                if isinstance(value, datetime):
                    value = value.isoformat()

                row.append(value or "")

            yield row

    async def get_export_summary(self, user_id: str) -> dict:
        """
        Get a summary of exportable data.

        Args:
            user_id: The user ID.

        Returns:
            Dictionary with counts and available filters.
        """
        success_doc = await success_applications_collection.find_one({"user_id": user_id})
        failed_doc = await failed_applications_collection.find_one({"user_id": user_id})

        success_count = len(success_doc.get("content", {})) if success_doc else 0
        failed_count = len(failed_doc.get("content", {})) if failed_doc else 0

        # Get unique portals
        portals = set()
        for doc in [success_doc, failed_doc]:
            if doc and "content" in doc:
                for job_data in doc["content"].values():
                    portal = job_data.get("portal")
                    if portal:
                        portals.add(portal)

        return {
            "total_applications": success_count + failed_count,
            "successful_applications": success_count,
            "failed_applications": failed_count,
            "available_portals": sorted(portals),
            "export_formats": ["csv", "excel"],
        }


# Global service instance
export_service = ExportService()
